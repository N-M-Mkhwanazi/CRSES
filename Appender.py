import sys
import json
import pandas as pd
import openpyxl
import logging
import difflib
import re
import os
from copy import copy
from PyQt6.QtWidgets import (QApplication, QMainWindow, QWidget, QLabel, QPushButton,
    QVBoxLayout, QHBoxLayout, QFileDialog, QMessageBox, QSizePolicy,
    QLineEdit, QProgressBar, QDoubleSpinBox, QSpacerItem, QTextEdit)
from PyQt6.QtCore import Qt, QThread, pyqtSignal, QObject
from openpyxl.utils import column_index_from_string
from openpyxl.cell.cell import Cell

# --- Path Configuration ---
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
MAP_PATH = os.path.join(BASE_DIR, "Map.json")

# --- Custom Log Handler for UI ---
class SignallingHandler(logging.Handler, QObject):
    log_signal = pyqtSignal(str)

    def __init__(self):
        logging.Handler.__init__(self)
        QObject.__init__(self)

    def emit(self, record):
        msg = self.format(record)
        self.log_signal.emit(msg)

# --- Helper Functions ---
def find_fuzzy_match(target_name, available_names, min_score=0.7):
    def _norm(name):
        return re.sub(r'\W+', '', str(name).lower())

    target_name_normalized = _norm(target_name)
    available_names_normalized = [_norm(n) for n in available_names]
    
    matches = difflib.get_close_matches(target_name_normalized, available_names_normalized, n=1, cutoff=min_score)
    
    if matches:
        matched_name_lower = matches[0]
        for name in available_names:
            if _norm(name) == matched_name_lower:
                return name
    return None

def percent_within(a, b, tol_percent):
    """Checks if a is within a certain percentage tolerance of b."""
    try:
        a_val = float(str(a).replace(',', '.'))
        b_val = float(str(b).replace(',', '.'))
        # Handle zero-division case
        if abs(b_val) < 1e-9:
            return abs(a_val) < 1e-9
        return abs(a_val - b_val) <= (tol_percent / 100.0) * abs(b_val)
    except (ValueError, TypeError):
        return False

def copy_cell_style(source_cell: Cell, target_cell: Cell):
    """Safely clones style attributes from a source cell to a target cell."""
    if source_cell.has_style:
        if source_cell.font:
            target_cell.font = copy(source_cell.font)
        if source_cell.fill:
            target_cell.fill = copy(source_cell.fill)
        if source_cell.border:
            target_cell.border = copy(source_cell.border)
        if source_cell.alignment:
            target_cell.alignment = copy(source_cell.alignment)
        if source_cell.protection:
            target_cell.protection = copy(source_cell.protection)
        target_cell.number_format = source_cell.number_format

# --- Background Worker ---
class AppendWorker(QThread):
    progress = pyqtSignal(int)
    status = pyqtSignal(str)
    finished = pyqtSignal(bool, str)

    def __init__(self, alpha_path, eskom_path, tolerance, mapping):
        super().__init__()
        self.alpha_path = alpha_path
        self.eskom_path = eskom_path
        self.tolerance = tolerance
        self.mapping = mapping

    def _get_alpha_column_index(self, mapping_tuple, alpha_stats_header_map):
        """Handle 2 or 3 element mapping tuples, including Excel column letters."""
        if len(mapping_tuple) == 3:
            alpha_specifier = mapping_tuple[2]
            
            # Handle Excel column letter specifier like "O" or "AN"
            if isinstance(alpha_specifier, str) and re.match(r'^[A-Z]+$', alpha_specifier, re.I):
                try:
                    return column_index_from_string(alpha_specifier)
                except ValueError:
                    logging.warning(f"Invalid Excel column letter specifier: '{alpha_specifier}'.")
                    return None
            elif isinstance(alpha_specifier, int):
                return alpha_specifier
            elif isinstance(alpha_specifier, str):
                alpha_col_indices = alpha_stats_header_map.get(alpha_specifier.strip().lower())
                if alpha_col_indices:
                    return alpha_col_indices[0]
                else:
                    logging.warning(f"Specified Alpha column '{alpha_specifier}' not found in header map.")
        
        # Fallback for 2-element or unresolved 3-element tuples
        alpha_column_name = mapping_tuple[1].strip().lower()
        alpha_col_indices = alpha_stats_header_map.get(alpha_column_name)
        if alpha_col_indices:
            return alpha_col_indices[0]
        return None

    def run(self):
        try:
            logging.info("Starting append process...")
            
            self.status.emit("Reading Eskom Data...")
            eskom_df = pd.read_excel(self.eskom_path)
            
            # Fuzzy match for date column
            date_col = find_fuzzy_match('Date Time Hour Beginning', eskom_df.columns.tolist())
            if not date_col:
                self.finished.emit(False, "Could not find date column in Eskom file (tried fuzzy match for 'Date Time Hour Beginning').")
                return
            
            logging.info(f"Found date column: '{date_col}'")
            
            # FIXED: Use lowercase 'h' for pandas 2.2+
            eskom_df[date_col] = pd.to_datetime(eskom_df[date_col]).dt.floor('h')
            
            self.progress.emit(20)
            
            self.status.emit("Loading Alpha File...")
            alpha_workbook = openpyxl.load_workbook(self.alpha_path)
            alpha_worksheet = alpha_workbook["Input"]  # Specific sheet name
            
            # FIXED: Load a second workbook with data_only=True for validation (reads calculated values)
            alpha_workbook_data_only = openpyxl.load_workbook(self.alpha_path, data_only=True)
            alpha_worksheet_data_only = alpha_workbook_data_only["Input"]
            
            # Build header map - handle duplicates
            alpha_stats_header_map = {}
            for cell in alpha_worksheet[1]:
                if cell.value:
                    key = str(cell.value).strip().lower()
                    col = cell.column
                    if isinstance(col, str):
                        col = column_index_from_string(col)
                    if key not in alpha_stats_header_map:
                        alpha_stats_header_map[key] = []
                    alpha_stats_header_map[key].append(col)
            
            # Find last data row
            last_data_row = alpha_worksheet.max_row
            while last_data_row > 1 and not any(alpha_worksheet.cell(row=last_data_row, column=c).value for c in range(1, alpha_worksheet.max_column + 1)):
                last_data_row -= 1
            
            self.progress.emit(35)
            
            # Validate date columns exist
            alpha_date_cols = [col.lower() for col in ["Year", "Month", "Day", "Hour"]]
            if not all(col in alpha_stats_header_map for col in alpha_date_cols):
                self.finished.emit(False, "Missing date/time columns (Year, Month, Day, Hour) in Alpha Stats.")
                return
            
            # Get last datetime from Alpha
            try:
                alpha_last_datetime = pd.to_datetime(
                    f"{alpha_worksheet.cell(row=last_data_row, column=alpha_stats_header_map['year'][0]).value}-"
                    f"{alpha_worksheet.cell(row=last_data_row, column=alpha_stats_header_map['month'][0]).value}-"
                    f"{alpha_worksheet.cell(row=last_data_row, column=alpha_stats_header_map['day'][0]).value} "
                    f"{alpha_worksheet.cell(row=last_data_row, column=alpha_stats_header_map['hour'][0]).value}:00:00"
                ).floor('h')
            except (ValueError, TypeError) as e:
                self.finished.emit(False, f"Error parsing date from Alpha Stats: {e}")
                return
            
            self.status.emit("Validating data...")
            
            # Find matching row in Eskom data
            matching_indices = eskom_df.index[eskom_df[date_col] == alpha_last_datetime]
            if matching_indices.empty:
                self.finished.emit(False, f"No matching timestamp '{alpha_last_datetime}' found in Eskom data.")
                return
            
            matching_row_index = matching_indices[0]
            
            # Validate values with tolerance
            validation_ok = True
            for mapping_tuple in self.mapping[1:]:
                eskom_column = mapping_tuple[0]
                alpha_column = mapping_tuple[1]
                
                # Fuzzy match for Eskom column
                eskom_column_found = find_fuzzy_match(eskom_column, eskom_df.columns.tolist())
                if not eskom_column_found:
                    logging.warning(f"Column '{eskom_column}' not found in Eskom file (fuzzy match failed), skipping validation.")
                    continue
                
                logging.info(f"Eskom column '{eskom_column}' matched to '{eskom_column_found}'")
                
                alpha_column_index = self._get_alpha_column_index(mapping_tuple, alpha_stats_header_map)
                if not alpha_column_index:
                    logging.warning(f"Column '{alpha_column}' missing in Alpha file, skipping validation.")
                    continue
                
                # Get the actual column letter for logging
                from openpyxl.utils import get_column_letter
                alpha_col_letter = get_column_letter(alpha_column_index)
                
                # FIXED: Read from data_only workbook to get calculated values instead of formulas
                alpha_value = alpha_worksheet_data_only.cell(row=last_data_row, column=alpha_column_index).value
                eskom_value = eskom_df.loc[matching_row_index, eskom_column_found]
                
                logging.info(f"Validating '{alpha_column}' (col {alpha_col_letter}): Alpha={alpha_value}, Eskom['{eskom_column_found}']={eskom_value}")
                
                if not percent_within(alpha_value, eskom_value, self.tolerance):
                    self.finished.emit(False, f"Mismatch in '{alpha_column}' (col {alpha_col_letter}):\nAlpha={alpha_value}\nEskom['{eskom_column_found}']={eskom_value}")
                    logging.error(f"Validation failed for '{alpha_column}'")
                    validation_ok = False
                    break
            
            if not validation_ok:
                return
            
            self.progress.emit(50)
            
            # Get rows to append
            pos = eskom_df.index.get_loc(matching_row_index)
            new_rows_to_append = eskom_df.iloc[pos + 1:].copy()
            
            if new_rows_to_append.empty:
                self.finished.emit(True, "Data is already up to date.")
                return
            
            self.status.emit(f"Appending {len(new_rows_to_append)} rows...")
            
            # Store last row cells for style copying
            last_row_cells = {}
            for col_idx in range(1, alpha_worksheet.max_column + 1):
                last_row_cells[col_idx] = alpha_worksheet.cell(row=last_data_row, column=col_idx)
            
            next_append_row = last_data_row + 1
            
            for row_num, (_, new_row) in enumerate(new_rows_to_append.iterrows()):
                eskom_datetime = pd.to_datetime(new_row[date_col])
                
                # Populate date columns
                alpha_date_cols_names = ["Year", "Week", "Month", "Day", "Hour"]
                alpha_date_values = [
                    int(eskom_datetime.year),
                    int(eskom_datetime.isocalendar().week),
                    int(eskom_datetime.month),
                    int(eskom_datetime.day),
                    int(eskom_datetime.hour)
                ]
                
                for alpha_col_name, alpha_value in zip(alpha_date_cols_names, alpha_date_values):
                    alpha_col_indices = alpha_stats_header_map.get(alpha_col_name.lower())
                    if alpha_col_indices:
                        col_idx = alpha_col_indices[0]
                        cell = alpha_worksheet.cell(row=next_append_row, column=col_idx)
                        cell.value = alpha_value
                        
                        # Copy style from last row but ensure number format is General for integers
                        if col_idx in last_row_cells:
                            copy_cell_style(last_row_cells[col_idx], cell)
                            # Override number format to ensure it displays as integer, not date
                            cell.number_format = '0'
                
                # Populate data columns
                for mapping_tuple in self.mapping[1:]:
                    eskom_column = mapping_tuple[0]
                    
                    # Fuzzy match for Eskom column
                    eskom_column_found = find_fuzzy_match(eskom_column, new_row.index.tolist())
                    if not eskom_column_found:
                        logging.warning(f"Column '{eskom_column}' not found in Eskom data (fuzzy match failed), skipping.")
                        continue
                    
                    alpha_col_index = self._get_alpha_column_index(mapping_tuple, alpha_stats_header_map)
                    if alpha_col_index:
                        value_to_append = new_row[eskom_column_found]
                        if isinstance(value_to_append, str):
                            value_to_append = value_to_append.replace(',', '.')
                        
                        new_cell = alpha_worksheet.cell(row=next_append_row, column=alpha_col_index)
                        new_cell.value = value_to_append
                        
                        # Copy style from last row
                        if alpha_col_index in last_row_cells:
                            copy_cell_style(last_row_cells[alpha_col_index], new_cell)
                
                next_append_row += 1
                
                # Update progress
                progress = 50 + int((row_num / len(new_rows_to_append)) * 40)
                self.progress.emit(progress)
            
            self.status.emit("Saving file...")
            alpha_workbook.save(self.alpha_path)
            self.progress.emit(100)
            
            self.finished.emit(True, f"Successfully appended {len(new_rows_to_append)} rows!")
            logging.info(f"Successfully appended {len(new_rows_to_append)} rows.")
            
        except Exception as e:
            logging.error(f"Error during append: {str(e)}", exc_info=True)
            self.finished.emit(False, f"Error: {str(e)}")

# --- Main Window ---
class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("CRSES - Alpha Appender Pro")
        self.resize(700, 600)
        
        self.alpha_file_path = ""
        self.eskom_file_path = ""
        self.mapping = []
        
        self.init_logging()
        self.load_mapping()
        self.init_ui()

    def init_logging(self):
        self.log_handler = SignallingHandler()
        self.log_handler.setFormatter(logging.Formatter('%(levelname)s: %(message)s'))
        logging.getLogger().addHandler(self.log_handler)
        logging.getLogger().setLevel(logging.INFO)

    def load_mapping(self):
        """Load Map.json at startup"""
        if not os.path.exists(MAP_PATH):
            QMessageBox.critical(self, "Error", f"Mapping file not found: {MAP_PATH}")
            sys.exit(1)
        
        try:
            with open(MAP_PATH, 'r') as f:
                self.mapping = json.load(f)
            logging.info(f"Loaded mapping with {len(self.mapping)} entries")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to load Map.json: {e}")
            sys.exit(1)

    def init_ui(self):
        central = QWidget()
        self.setCentralWidget(central)
        layout = QVBoxLayout(central)

        # Header
        logo_layout = QHBoxLayout()
        logo_layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
        
        sun_motif = QLabel("⬤")
        sun_motif.setStyleSheet("color: #FFB81C; font-size: 40px; font-weight: bold;")
        logo_layout.addWidget(sun_motif)
        
        title = QLabel("CRSES")
        title.setStyleSheet("color: #7A003C; font-size: 36px; font-weight: bold;")
        logo_layout.addWidget(title)
        
        layout.addLayout(logo_layout)
        
        subtitle = QLabel("Excel Data Appender")
        subtitle.setAlignment(Qt.AlignmentFlag.AlignCenter)
        subtitle.setStyleSheet("color: #4D4D4D; font-size: 18px;")
        layout.addWidget(subtitle)

        # Files
        self.alpha_input = QLineEdit()
        self.alpha_input.setPlaceholderText("Select Alpha Stats file...")
        self.alpha_input.setReadOnly(True)
        btn_alpha = QPushButton("Browse Alpha File")
        btn_alpha.setStyleSheet("QPushButton { color: #4D4D4D; border-radius: 10px; }"
                                "QPushButton:hover { background-color: #808080; }"
                                "QPushButton:pressed { background-color: #CC9417; }")
        btn_alpha.clicked.connect(lambda: self.get_file(True))

        self.eskom_input = QLineEdit()
        self.eskom_input.setPlaceholderText("Select Eskom Data file...")
        self.eskom_input.setReadOnly(True)
        btn_eskom = QPushButton("Browse Eskom File")
        btn_eskom.setStyleSheet("QPushButton { color: #4D4D4D; border-radius: 10px; }"
                                "QPushButton:hover { background-color: #808080; }"
                                "QPushButton:pressed { background-color: #CC9417; }")
        btn_eskom.clicked.connect(lambda: self.get_file(False))

        for row in [(self.alpha_input, btn_alpha), (self.eskom_input, btn_eskom)]:
            h = QHBoxLayout()
            h.addWidget(row[0])
            h.addWidget(row[1])
            layout.addLayout(h)

        # Config
        cfg_layout = QHBoxLayout()
        cfg_layout.addWidget(QLabel("Tolerance %:"))
        self.tolerance_box = QDoubleSpinBox()
        self.tolerance_box.setRange(0.0, 100.0)
        self.tolerance_box.setValue(5.0)
        self.tolerance_box.setSingleStep(1.0)
        cfg_layout.addWidget(self.tolerance_box)
        cfg_layout.addStretch()
        layout.addLayout(cfg_layout)

        # Progress & Status
        self.status_label = QLabel("Ready: Please select your files.")
        self.status_label.setStyleSheet("background-color: #F8F8F8; color: #4D4D4D; padding: 10px; border-radius: 8px; border: 1px solid #E0E0E0;")
        layout.addWidget(self.status_label)
        
        self.progress = QProgressBar()
        self.progress.setStyleSheet("QProgressBar::chunk { background-color: #FFB81C; border-radius: 5px; }")
        layout.addWidget(self.progress)

        # Log Console
        self.console = QTextEdit()
        self.console.setReadOnly(True)
        self.console.setStyleSheet("background: #F0F0F0; font-family: Consolas; font-size: 11px;")
        layout.addWidget(QLabel("Activity Log:"))
        layout.addWidget(self.console)
        self.log_handler.log_signal.connect(self.console.append)

        # Run Button
        self.run_btn = QPushButton("Run Appender")
        self.run_btn.setFixedSize(200, 50)
        self.run_btn.setStyleSheet(
            "QPushButton { background-color: #7A003C; color: white; border-radius: 10px; font-weight: bold; }"
            "QPushButton:hover { background-color: #6C0035; }"
            "QPushButton:pressed { background-color: #5E002D; }"
            "QPushButton:disabled { background-color: #cccccc; color: #666666; }"
        )
        self.run_btn.clicked.connect(self.start_process)
        self.run_btn.setEnabled(False)
        layout.addWidget(self.run_btn, alignment=Qt.AlignmentFlag.AlignCenter)

    def get_file(self, is_alpha):
        path, _ = QFileDialog.getOpenFileName(self, "Select Excel File", "", "Excel Files (*.xlsx);;All Files (*)")
        if path:
            if is_alpha:
                self.alpha_file_path = path
                self.alpha_input.setText(path)
            else:
                self.eskom_file_path = path
                self.eskom_input.setText(path)
            
            if self.alpha_file_path and self.eskom_file_path:
                self.run_btn.setEnabled(True)
                self.status_label.setText("Files selected. Click 'Run Appender' to start.")

    def start_process(self):
        if not self.alpha_file_path or not self.eskom_file_path:
            QMessageBox.warning(self, "Missing Files", "Please select both files.")
            return

        self.run_btn.setEnabled(False)
        self.progress.setValue(0)
        
        self.worker = AppendWorker(
            self.alpha_file_path,
            self.eskom_file_path,
            self.tolerance_box.value(),
            self.mapping
        )
        self.worker.progress.connect(self.progress.setValue)
        self.worker.status.connect(self.status_label.setText)
        self.worker.finished.connect(self.on_finished)
        self.worker.start()

    def on_finished(self, success, message):
        self.run_btn.setEnabled(True)
        if success:
            QMessageBox.information(self, "Success", message)
            self.status_label.setText("Done! Data appended successfully.")
        else:
            QMessageBox.critical(self, "Error", message)
            self.status_label.setText("Error: Process failed.")

if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec())
